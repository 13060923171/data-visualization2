import datetime

import openpyxl
import pyecharts.options as opts
from pyecharts.charts import Map, Line
from pyecharts.globals import CurrentConfig, NotebookType

CurrentConfig.NOTEBOOK_TYPE = NotebookType.JUPYTER_LAB
"""国内外疫情可视化 pyecharts 新增确诊"""
def china_map ():
    wb = openpyxl.load_workbook('demo/data/国内疫情.xlsx') # 获取已有的xlsx文件
    ws_data = wb['Sheet1']  # 获取文件中中国省份疫情数据表
    ws_data.delete_rows(1)  # 删除第一行
    province = []  # 省份
    now_curconfirm = []  # 现确诊
    #获取表数据
    for data in ws_data.values:
        province.append(data[0])
        now_curconfirm.append(data[1])
    #绘制地图
    m = Map()
    m.add("各省现有确诊："
          , [list(z) for z in zip(province, now_curconfirm)
    ],
          maptype="china",
          is_map_symbol_show=False)
    # 设置全局属性
    m.set_global_opts(
        # 标题
        title_opts=opts.TitleOpts(title="COVID-19中国现有地区现有确诊人数地图"+"\n"+str(datetime.date.today())),
        # 分段型数据，自定义分段
        visualmap_opts=opts.VisualMapOpts(
            is_piecewise=True,
            pieces=[
                {
                    "min": 1000,
                    "label": '>1000',
                    "color": "#893448"
                },  # 不指定 max，表示 max 为无限大
                {
                    "min": 500,
                    "max": 999,
                    "label": '500-1000',
                    "color": "#ff585e"
                },
                {
                    "min": 100,
                    "max": 499,
                    "label": '100-500',
                    "color": "#fb8146"
                },
                {
                    "min": 50,
                    "max": 99,
                    "label": '50-99',
                    "color": "#ffA500"
                },
                {
                    "min": 10,
                    "max": 49,
                    "label": '10-49',
                    "color": "#ffb248"
                },
                {
                    "min": 1,
                    "max": 9,
                    "label": '1-9',
                    "color": "#f59e83"
                },
                {
                    "max": 1,
                    "label": '0',
                    "color": "#fdebcf"
                }
            ]))
    # 设置坐标属性，显示名称
    m.set_series_opts(
        label_opts=opts.LabelOpts(is_show=True)
    )

    print("中国地图已绘制")
    m.render('demo/result/中国各省市疫情地图.html')
    return m

def china_linechart():
    wb = openpyxl.load_workbook('demo/data/国内历史疫情.xlsx')  # 获取已有的xlsx文件
    ws_data = wb['Sheet1']  # 获取文件中中国省份疫情数据表
    ws_data.delete_rows(1)  # 删除第一行
    date = []  # 时间
    heal = []  # 现确诊
    curconfirm = [] #累计确诊
    dead = [] #死亡
    # 获取表数据
    for data in ws_data.values:
        curconfirm.append(data[0])
        date.append(data[1])
        heal.append(data[2])
        dead.append(data[3])
    #绘制折线图
    line_chart = Line()
    line_chart.add_xaxis(date)
    line_chart.add_yaxis("累计确认", curconfirm)
    line_chart.add_yaxis("死亡", dead)
    line_chart.add_yaxis("治愈",heal)
    line_chart.set_global_opts(
        title_opts=opts.TitleOpts(title="新冠疫情累计病例与治愈病例、死亡曲线")
         )
    #不显示标签数据
    line_chart.set_series_opts(label_opts=opts.LabelOpts(is_show=False))
    print("中国曲线图已绘制")
    line_chart.render('demo/result/中国新冠疫情曲线-{}.html'.format(datetime.date.today()))
    return line_chart

def india_linechart():
    wb = openpyxl.load_workbook('demo/data/印度历史疫情.xlsx')  # 获取已有的xlsx文件
    ws_data = wb['Sheet1']  # 获取文件中中国省份疫情数据表
    ws_data.delete_rows(1)  # 删除第一行
    date = []  # 时间
    heal = []  # 现确诊
    curconfirm = []  # 累计确诊
    confirm_add = [] #新增确诊
    dead = []  # 死亡
    # 获取表数据
    for data in ws_data.values:
        curconfirm.append(data[0])
        date.append(data[1])
        heal.append(data[3])
        dead.append(data[4])
        confirm_add.append(data[2])
    # 绘制折线图
    line_chart = Line()
    line_chart.add_xaxis(date)
    line_chart.add_yaxis("累计确认", curconfirm)
    line_chart.add_yaxis("死亡", dead)
    line_chart.add_yaxis("治愈", heal)
    line_chart.add_yaxis("新增确认", confirm_add)
    line_chart.set_global_opts(
        title_opts=opts.TitleOpts(title="印度新冠疫情曲线")
    )
    # 不显示标签数据
    line_chart.set_series_opts(label_opts=opts.LabelOpts(is_show=False))
    print("印度曲线图已绘制")
    line_chart.render('demo/result/印度新冠疫情曲线-{}.html'.format(datetime.date.today()))
    return line_chart

if __name__ == "__main__":
    #测试
    china_map ()
    china_linechart()
    india_linechart()

